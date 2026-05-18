"""Tests for simple_org_chart.exports – format_export_filters."""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from simple_org_chart.exports import add_metadata_sheet, format_export_filters


class TestFormatExportFilters:
    def test_scope_only(self):
        result = format_export_filters('orgChart')
        assert result == 'scope=orgChart'

    def test_scope_all(self):
        result = format_export_filters('all')
        assert result == 'scope=all'

    def test_with_toggles(self):
        toggles = {
            'include_user_mailboxes': True,
            'include_shared_mailboxes': False,
            'include_room_equipment_mailboxes': False,
            'include_enabled': True,
            'include_disabled': False,
            'include_licensed': True,
            'include_unlicensed': True,
            'include_members': True,
            'include_guests': False,
        }
        result = format_export_filters('orgChart', toggles=toggles)
        assert 'scope=orgChart' in result
        assert 'userMailboxes=yes' in result
        assert 'sharedMailboxes=no' in result
        assert 'enabled=yes' in result
        assert 'disabled=no' in result
        assert 'guests=no' in result

    def test_with_tagpicker_titles(self):
        tp = {
            'filter_titles': ['Engineer', 'Manager'],
            'filter_titles_mode': 'include',
            'filter_departments': None,
            'filter_departments_mode': 'exclude',
            'filter_countries': None,
            'filter_countries_mode': 'exclude',
        }
        result = format_export_filters('all', tp=tp)
        assert 'scope=all' in result
        assert 'titles(include): Engineer, Manager' in result
        assert 'departments' not in result
        assert 'countries' not in result

    def test_with_tagpicker_all_three(self):
        tp = {
            'filter_titles': ['Dev'],
            'filter_titles_mode': 'exclude',
            'filter_departments': ['Engineering', 'Sales'],
            'filter_departments_mode': 'include',
            'filter_countries': ['US'],
            'filter_countries_mode': 'exclude',
        }
        result = format_export_filters('orgChart', tp=tp)
        assert 'titles(exclude): Dev' in result
        assert 'departments(include): Engineering, Sales' in result
        assert 'countries(exclude): US' in result

    def test_with_toggles_and_tagpicker(self):
        toggles = {
            'include_enabled': True,
            'include_disabled': False,
        }
        tp = {
            'filter_titles': ['VP'],
            'filter_titles_mode': 'include',
            'filter_departments': None,
            'filter_departments_mode': 'exclude',
            'filter_countries': None,
            'filter_countries_mode': 'exclude',
        }
        result = format_export_filters('orgChart', toggles=toggles, tp=tp)
        assert 'scope=orgChart' in result
        assert 'enabled=yes' in result
        assert 'disabled=no' in result
        assert 'titles(include): VP' in result

    def test_empty_tagpicker_values_ignored(self):
        tp = {
            'filter_titles': [],
            'filter_titles_mode': 'exclude',
            'filter_departments': None,
            'filter_departments_mode': 'exclude',
            'filter_countries': None,
            'filter_countries_mode': 'exclude',
        }
        result = format_export_filters('all', tp=tp)
        assert result == 'scope=all'

    def test_none_toggles_and_tp(self):
        result = format_export_filters('orgChart', toggles=None, tp=None)
        assert result == 'scope=orgChart'


class TestAddMetadataSheet:
    def test_excludes_exported_by_when_empty(self):
        wb = Workbook()
        add_metadata_sheet(
            wb,
            filename='report.xlsx',
            sheet_title='Data',
            item_count=3,
            data_export_option='scope=all',
        )
        labels = [cell.value for cell in wb['Metadata']['A'] if cell.value]
        assert 'Exported by' not in labels

    def test_includes_exported_by_when_provided(self):
        wb = Workbook()
        add_metadata_sheet(
            wb,
            filename='report.xlsx',
            sheet_title='Data',
            item_count=3,
            data_export_option='scope=all',
            exported_by='alice@example.com',
        )
        metadata_ws = wb['Metadata']
        rows = {
            metadata_ws.cell(row=row, column=1).value: metadata_ws.cell(row=row, column=2).value
            for row in range(4, metadata_ws.max_row + 1)
        }
        assert rows.get('Exported by') == 'alice@example.com'
