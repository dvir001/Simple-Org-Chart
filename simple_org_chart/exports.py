"""Export utilities for SimpleOrgChart."""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def format_hire_date(date_string: Optional[str]) -> str:
    """Format an ISO date string to a readable format."""
    if not date_string:
        return ''
    try:
        dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return date_string


def _resolve_logo_path() -> Optional[str]:
    """Return the filesystem path to the configured logo image, or *None*."""
    from simple_org_chart.config import BASE_DIR, DATA_DIR
    try:
        from simple_org_chart.settings import load_settings
        logo_web_path = load_settings().get('logoPath', '/static/icon.png')
    except Exception:
        logo_web_path = '/static/icon.png'

    # Custom logo: /static/icon_custom_<hash>.png  →  data/icon_custom_<hash>.png
    basename = os.path.basename(logo_web_path or '')
    if basename.startswith('icon_custom_'):
        candidate = DATA_DIR / basename
    else:
        candidate = BASE_DIR / 'static' / 'icon.png'

    if candidate.is_file():
        return str(candidate)
    return None


def add_metadata_sheet(
    wb,
    *,
    filename: str,
    sheet_title: str,
    item_count: int = 0,
    data_export_option: str = '',
    exported_by: str = '',
):
    """Append a **Metadata** sheet to an openpyxl Workbook.

    Mirrors the layout used by Microsoft Purview DSPM exports.
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.drawing.image import Image as XlImage

    ws = wb.create_sheet(title='Metadata')

    label_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    value_alignment = Alignment(horizontal='left')

    # Logo image in A1 (if available)
    logo_path = _resolve_logo_path()
    logo_col_offset = 1  # start text in column A by default
    if logo_path:
        try:
            img = XlImage(logo_path)
            ws.add_image(img, 'A1')
            logo_col_offset = 2  # push title text to column B
        except Exception:
            logo_col_offset = 1

    # Title text next to logo
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    title_cell = ws.cell(row=1, column=logo_col_offset, value='Simple Org Chart')
    title_cell.font = title_font
    title_cell.alignment = Alignment(vertical='center')

    subtitle_cell = ws.cell(row=2, column=logo_col_offset, value='Export Report')
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(vertical='center')

    # Metadata rows start at row 4
    rows = [
        ('Generated on', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')),
        ('Format', 'XLSX'),
        ('File name', filename),
        ('Data sheet', sheet_title),
        ('Data export option', data_export_option),
        ('Items exported', item_count),
        ('Exported by', exported_by),
    ]

    for idx, (label, value) in enumerate(rows, start=4):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = label_font
        value_cell = ws.cell(row=idx, column=2, value=value)
        value_cell.alignment = value_alignment

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 40


def format_export_filters(scope, toggles=None, tp=None):
    """Build a descriptive data_export_option string with all active filters."""
    parts = [f"scope={scope}"]

    if toggles:
        toggle_labels = {
            'include_user_mailboxes': 'userMailboxes',
            'include_shared_mailboxes': 'sharedMailboxes',
            'include_room_equipment_mailboxes': 'roomEquipmentMailboxes',
            'include_enabled': 'enabled',
            'include_disabled': 'disabled',
            'include_licensed': 'licensed',
            'include_unlicensed': 'unlicensed',
            'include_members': 'members',
            'include_guests': 'guests',
        }
        for key, label in toggle_labels.items():
            val = toggles.get(key)
            if val is not None:
                parts.append(f"{label}={'yes' if val else 'no'}")

    if tp:
        for filter_key, mode_key, label in [
            ('filter_titles', 'filter_titles_mode', 'titles'),
            ('filter_departments', 'filter_departments_mode', 'departments'),
            ('filter_countries', 'filter_countries_mode', 'countries'),
        ]:
            values = tp.get(filter_key)
            if values:
                mode = tp.get(mode_key, 'exclude')
                parts.append(f"{label}({mode}): {', '.join(values)}")

    return ', '.join(parts)


__all__ = [
    'format_hire_date',
    'add_metadata_sheet',
    'format_export_filters',
]
