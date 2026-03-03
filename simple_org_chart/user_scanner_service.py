"""User Scanner integration service for SimpleOrgChart.

Downloads user-scanner from PyPI on demand and runs email/username
OSINT scans via its library API.  The package is installed into
``repositories/user-scanner`` which is git-ignored so source control
stays clean.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import subprocess
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

import simple_org_chart.config as app_config

logger = logging.getLogger(__name__)

# Suppress noisy per-request HTTP logs from httpx / httpcore so they don't
# flood docker logs during scans.
for _noisy in ('httpx', 'httpcore', 'httpcore.http11', 'httpcore.connection'):
    logging.getLogger(_noisy).setLevel(logging.WARNING)

# Delay between employees in a full scan (seconds) to avoid triggering
# rate-limits (HTTP 429) on third-party services.
_SCAN_DELAY_SECONDS = 1.5

REPO_DIR = app_config.BASE_DIR / "repositories" / "user-scanner"
USER_SCANNER_CACHE = app_config.DATA_DIR / "user_scanner_results.json"
USER_SCANNER_HISTORY = app_config.DATA_DIR / "user_scanner_history.json"
USER_SCANNER_XLSX_DIR = app_config.DATA_DIR / "user_scanner_exports"
MAX_SCAN_HISTORY = 5
PYPI_PACKAGE_NAME = "user-scanner"


# ---------------------------------------------------------------------------
# Installation helpers
# ---------------------------------------------------------------------------

def is_installed() -> bool:
    """Return True when the user-scanner package exists locally."""
    return (REPO_DIR / "user_scanner").is_dir()


def install() -> bool:
    """Download user-scanner from PyPI into *REPO_DIR*.

    Returns True on success, False on failure.
    """
    REPO_DIR.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.check_call(
            [
                sys.executable, "-m", "pip", "install",
                "user-scanner",
                "--target", str(REPO_DIR),
                "--upgrade",
                "--quiet",
            ],
            timeout=120,
        )
        logger.info("user-scanner installed successfully to %s", REPO_DIR)
        return True
    except Exception as exc:
        logger.error("Failed to install user-scanner: %s", exc)
        return False


def get_version() -> Optional[str]:
    """Return the installed user-scanner version, or None."""
    if not is_installed():
        return None
    # Primary: read version from dist-info METADATA in the target directory
    try:
        for item in REPO_DIR.iterdir():
            if (
                item.is_dir()
                and item.name.endswith(".dist-info")
                and item.name.lower().startswith("user_scanner-")
            ):
                metadata_file = item / "METADATA"
                if metadata_file.exists():
                    for line in metadata_file.read_text(encoding="utf-8").splitlines():
                        if line.startswith("Version:"):
                            return line.split(":", 1)[1].strip()
    except Exception:
        pass
    # Fallback: try the package attribute
    try:
        _ensure_on_path()
        import importlib
        mod = importlib.import_module("user_scanner")
        ver = getattr(mod, "__version__", None)
        if ver:
            return str(ver)
    except Exception:
        pass
    return None


def _ensure_on_path() -> None:
    """Make sure the local install folder is on ``sys.path``."""
    repo_str = str(REPO_DIR)
    if repo_str not in sys.path:
        sys.path.insert(0, repo_str)


def get_latest_pypi_version() -> Optional[str]:
    """Query PyPI for the latest release version of user-scanner."""
    try:
        resp = requests.get(
            f"https://pypi.org/pypi/{PYPI_PACKAGE_NAME}/json",
            timeout=10,
        )
        if resp.ok:
            return resp.json().get("info", {}).get("version")
    except Exception as exc:
        logger.warning("Failed to check PyPI for user-scanner updates: %s", exc)
    return None


def check_for_update() -> Dict[str, Any]:
    """Compare installed version against PyPI and return status dict."""
    installed = is_installed()
    current = get_version() if installed else None
    latest = get_latest_pypi_version()
    update_available = False
    if current and latest:
        update_available = _version_tuple(latest) > _version_tuple(current)
    return {
        "installed": installed,
        "currentVersion": current,
        "latestVersion": latest,
        "updateAvailable": update_available,
    }


def _version_tuple(v: str) -> tuple:
    """Convert a version string like '1.3.3' to a comparable tuple."""
    parts = []
    for segment in v.split("."):
        try:
            parts.append(int(segment))
        except ValueError:
            parts.append(segment)
    return tuple(parts)


# ---------------------------------------------------------------------------
# Scan helpers – thin wrappers around the library async API
# ---------------------------------------------------------------------------

def _run_async(coro):
    """Run an async coroutine from synchronous code."""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None

    if loop and loop.is_running():
        # We are inside an existing event-loop (e.g. Quart / Gunicorn).
        # Spin up a fresh loop in the current thread context.
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            return pool.submit(asyncio.run, coro).result()
    return asyncio.run(coro)


def _result_to_dict(result) -> Dict[str, Any]:
    """Normalise a ``Result`` object into a plain dict."""
    return {
        "site_name": getattr(result, "site_name", None) or "",
        "category": getattr(result, "category", None) or "",
        "status": result.status.to_label(getattr(result, "is_email", False)),
        "url": getattr(result, "url", "") or "",
        "reason": result.get_reason() if hasattr(result, "get_reason") else "",
        "is_email": getattr(result, "is_email", False),
    }


def get_available_sites(is_email: bool = True) -> List[str]:
    """Return sorted list of all site names the scanner can check."""
    if not is_installed():
        return []
    try:
        _ensure_on_path()
        from user_scanner.core.helpers import load_categories, load_modules, get_site_name  # type: ignore[import-untyped]

        names: List[str] = []
        categories = load_categories(is_email=is_email)
        for cat_path in categories.values():
            for mod in load_modules(cat_path):
                name = get_site_name(mod)
                if name:
                    names.append(name)
        names.sort(key=lambda s: s.lower())
        return names
    except Exception as exc:
        logger.warning("Failed to list available sites: %s", exc)
        return []


def get_loud_sites(is_email: bool = True) -> List[str]:
    """Return sorted list of site names that are considered 'loud'."""
    if not is_installed():
        return []
    try:
        _ensure_on_path()
        from user_scanner.core.helpers import (
            load_categories, load_modules, get_site_name, is_loud,
        )

        names: List[str] = []
        categories = load_categories(is_email=is_email)
        for cat_path in categories.values():
            for mod in load_modules(cat_path):
                name = get_site_name(mod)
                if name and is_loud(name, is_email=is_email):
                    names.append(name)
        names.sort(key=lambda s: s.lower())
        return names
    except Exception as exc:
        logger.warning("Failed to list loud sites: %s", exc)
        return []


def get_available_categories(is_email: bool = True) -> List[str]:
    """Return sorted list of scanner category names (e.g. social, gaming)."""
    if not is_installed():
        return []
    try:
        _ensure_on_path()
        from user_scanner.core.helpers import load_categories  # type: ignore[import-untyped]

        cats = sorted(load_categories(is_email=is_email).keys(),
                      key=lambda s: s.lower())
        return cats
    except Exception as exc:
        logger.warning("Failed to list scanner categories: %s", exc)
        return []


def scan_email(
    email: str,
    *,
    sites: Optional[List[str]] = None,
    categories: Optional[List[str]] = None,
    allow_loud: bool = False,
    only_found: bool = False,
) -> List[Dict[str, Any]]:
    """Run an email scan and return a list of result dicts.

    Parameters
    ----------
    sites : list[str] | None
        If given, scan only these site display-names.
    categories : list[str] | None
        If given, scan only sites in these categories.
    allow_loud : bool
        Include "loud" sites that may notify the target.
    only_found : bool
        Filter results to only registered/found entries.
    """
    _ensure_on_path()

    if sites or categories or not allow_loud:
        # Use the filtered path whenever any constraint is active
        return _scan_filtered(
            email, is_email=True,
            site_names=sites, category_names=categories,
            allow_loud=allow_loud, only_found=only_found,
        )

    from user_scanner.core.engine import check_all  # type: ignore[import-untyped]
    raw_results = _run_async(check_all(email, is_email=True))
    results = [_result_to_dict(r) for r in raw_results]
    if only_found:
        results = [r for r in results if r["status"] in ("Registered", "Found")]
    return results


def _scan_filtered(
    target: str,
    is_email: bool,
    *,
    site_names: Optional[List[str]] = None,
    category_names: Optional[List[str]] = None,
    allow_loud: bool = False,
    only_found: bool = False,
) -> List[Dict[str, Any]]:
    """Scan a filtered subset of sites.

    Filters applied (all optional, combined with AND):
    * *site_names* – only sites whose display-name is in this list.
    * *category_names* – only sites in these categories.
    * *allow_loud* – if False (default) loud modules are skipped.
    """
    from user_scanner.core.engine import check  # type: ignore[import-untyped]
    from user_scanner.core.helpers import (  # type: ignore[import-untyped]
        load_categories, load_modules, get_site_name, is_loud,
    )

    site_lower = {s.lower() for s in site_names} if site_names else None
    cat_lower = {c.lower() for c in category_names} if category_names else None

    all_categories = load_categories(is_email=is_email)
    tasks = []
    for cat_name, cat_path in all_categories.items():
        if cat_lower and cat_name.lower() not in cat_lower:
            continue
        for mod in load_modules(cat_path):
            name = get_site_name(mod)
            if site_lower and name.lower() not in site_lower:
                continue
            if not allow_loud and is_loud(name, is_email=is_email):
                continue
            tasks.append(check(mod, target))

    if not tasks:
        return []

    async def _gather():
        return list(await asyncio.gather(*tasks))

    raw_results = _run_async(_gather())
    results = [_result_to_dict(r) for r in raw_results]
    if only_found:
        results = [r for r in results if r["status"] in ("Registered", "Found")]
    return results


def scan_username(username: str) -> List[Dict[str, Any]]:
    """Run a full username scan and return a list of result dicts."""
    _ensure_on_path()
    from user_scanner.core.engine import check_all  # type: ignore[import-untyped]

    raw_results = _run_async(check_all(username, is_email=False))
    return [_result_to_dict(r) for r in raw_results]


def scan_user(
    email: str,
    *,
    sites: Optional[List[str]] = None,
    categories: Optional[List[str]] = None,
    allow_loud: bool = False,
    only_found: bool = False,
) -> Dict[str, Any]:
    """Scan a user by their Office 365 email address.

    Returns a dict with ``email_results`` and metadata.
    """
    if not is_installed():
        raise RuntimeError("user-scanner is not installed.  Enable it in settings first.")

    timestamp = datetime.now(timezone.utc).isoformat()
    payload: Dict[str, Any] = {
        "email": email,
        "scannedAt": timestamp,
        "email_results": [],
    }

    try:
        payload["email_results"] = scan_email(
            email,
            sites=sites,
            categories=categories,
            allow_loud=allow_loud,
            only_found=only_found,
        )
    except Exception as exc:
        logger.error("Email scan failed for %s: %s", email, exc)
        payload["email_error"] = str(exc)

    return payload


# ---------------------------------------------------------------------------
# Full-org scan (runs in background thread, results cached to disk)
# ---------------------------------------------------------------------------

def run_full_scan(
    employees: List[Dict[str, Any]],
    *,
    sites: Optional[List[str]] = None,
    categories: Optional[List[str]] = None,
    allow_loud: bool = False,
    only_found: bool = False,
    cancel_event: Optional[Any] = None,
    cancel_check: Optional[Any] = None,
    progress_callback: Optional[Any] = None,
) -> Dict[str, Any]:
    """Scan every employee by email.  Returns summary dict and caches to disk.

    Cancellation is checked via *cancel_event* (threading.Event) **or**
    *cancel_check* (callable returning bool).  Either mechanism will cause
    the loop to stop early and return partial results with ``_cancelled=True``.

    *progress_callback*, when supplied, is called with a dict describing
    each step so the caller can expose live progress to the UI.
    """
    if not is_installed():
        raise RuntimeError("user-scanner is not installed.")

    def _is_cancelled() -> bool:
        if cancel_event is not None and cancel_event.is_set():
            return True
        if cancel_check is not None:
            try:
                return bool(cancel_check())
            except Exception:
                return False
        return False

    def _progress(kind: str, **kwargs: Any) -> None:
        if progress_callback is not None:
            try:
                progress_callback({'kind': kind, **kwargs})
            except Exception:
                pass  # never let callback errors break the scan

    timestamp = datetime.now(timezone.utc).isoformat()
    results: List[Dict[str, Any]] = []
    cancelled = False

    # Count scannable (have email) employees for accurate total
    scannable = [e for e in employees if e.get('email') or e.get('mail')]
    total = len(scannable)

    _progress('start', total=total)

    for idx, emp in enumerate(scannable):
        # Check for cancellation before each employee
        if _is_cancelled():
            cancelled = True
            _progress('cancelled', scanned=len(results), total=total)
            break

        # Rate-limit: pause between employees to avoid 429s
        if idx > 0:
            time.sleep(_SCAN_DELAY_SECONDS)

        email = emp.get("email") or emp.get("mail")
        name = emp.get("name") or emp.get("displayName") or ""

        _progress('scanning', current=idx + 1, total=total, email=email, name=name)

        try:
            scan_results = scan_email(
                email,
                sites=sites,
                categories=categories,
                allow_loud=allow_loud,
                only_found=only_found,
            )
            registered_count = sum(
                1 for r in scan_results if r.get("status") == "Registered"
            )
            results.append({
                "name": name,
                "email": email,
                "totalChecked": len(scan_results),
                "registeredCount": registered_count,
                "results": scan_results,
            })
            _progress('scanned', current=idx + 1, total=total,
                      email=email, name=name,
                      registered=registered_count,
                      checked=len(scan_results))
        except Exception as exc:
            logger.warning("Scan failed for employee '%s': %s", name or "<unknown>", exc)
            results.append({
                "name": name,
                "email": email,
                "error": str(exc),
                "totalChecked": 0,
                "registeredCount": 0,
                "results": [],
            })
            _progress('error', current=idx + 1, total=total,
                      email=email, name=name, error=str(exc))

    payload = {
        "scannedAt": timestamp,
        "totalEmployees": len(results),
        "records": results,
        "_cancelled": cancelled,
    }

    # Persist to cache
    try:
        USER_SCANNER_CACHE.parent.mkdir(parents=True, exist_ok=True)
        with USER_SCANNER_CACHE.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2)
        logger.info("Full user-scanner results cached (%d employees)", len(results))
    except Exception as exc:
        logger.error("Failed to cache user-scanner results: %s", exc)

    # Generate XLSX and record in history
    try:
        xlsx_path = generate_xlsx(payload)
        history_entry = record_scan_run(payload, xlsx_path)
        payload["_scanId"] = history_entry["scanId"]
        payload["_xlsxFile"] = history_entry["xlsxFile"]
    except Exception as exc:
        logger.error("Failed to generate XLSX / record history: %s", exc)

    return payload


def load_cached_full_scan() -> Optional[Dict[str, Any]]:
    """Load the most recent full-scan results from disk."""
    if not USER_SCANNER_CACHE.exists():
        return None
    try:
        with USER_SCANNER_CACHE.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception as exc:
        logger.error("Failed to load cached user-scanner results: %s", exc)
        return None


# ---------------------------------------------------------------------------
# XLSX generation – one worksheet per scanner site
# ---------------------------------------------------------------------------

def _safe_sheet_name(name: str) -> str:
    """Sanitise a string so it can be used as an Excel sheet name."""
    clean = re.sub(r'[\[\]:*?/\\]', '_', name or 'Unknown')
    return clean[:31] or 'Sheet'


def generate_xlsx(scan_result: Dict[str, Any]) -> Path:
    """Create an XLSX workbook from a full-scan result.

    * **Summary** sheet – one row per employee with totals
    * One sheet per scanner *site_name* – rows for each employee found there

    Returns the path to the generated file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')

    summary_headers = ['Name', 'Email', 'Sites Checked', 'Registered', 'Error']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    site_data: Dict[str, List[Dict[str, str]]] = {}

    for idx, rec in enumerate(scan_result.get('records', []), start=2):
        ws_summary.cell(row=idx, column=1, value=rec.get('name', ''))
        ws_summary.cell(row=idx, column=2, value=rec.get('email', ''))
        ws_summary.cell(row=idx, column=3, value=rec.get('totalChecked', 0))
        ws_summary.cell(row=idx, column=4, value=rec.get('registeredCount', 0))
        ws_summary.cell(row=idx, column=5, value=rec.get('error', ''))

        for r in rec.get('results', []):
            site = r.get('site_name') or 'Unknown'
            status = (r.get('status') or '').lower()
            if status in ('registered', 'found'):
                site_data.setdefault(site, []).append({
                    'name': rec.get('name', ''),
                    'email': rec.get('email', ''),
                    'category': r.get('category', ''),
                    'status': r.get('status', ''),
                    'url': r.get('url', ''),
                    'reason': r.get('reason', ''),
                })

    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[chr(64 + col)].width = 28

    # --- Per-site sheets ---
    site_headers = ['Name', 'Email', 'Category', 'Status', 'URL', 'Reason']
    for site_name in sorted(site_data.keys(), key=lambda s: s.lower()):
        sheet_title = _safe_sheet_name(site_name)
        existing = {ws.title for ws in wb.worksheets}
        suffix = 1
        orig_title = sheet_title
        while sheet_title in existing:
            sheet_title = f'{orig_title[:28]}_{suffix}'
            suffix += 1

        ws = wb.create_sheet(title=sheet_title)
        for col, h in enumerate(site_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, entry in enumerate(site_data[site_name], start=2):
            ws.cell(row=row_idx, column=1, value=entry['name'])
            ws.cell(row=row_idx, column=2, value=entry['email'])
            ws.cell(row=row_idx, column=3, value=entry['category'])
            ws.cell(row=row_idx, column=4, value=entry['status'])
            ws.cell(row=row_idx, column=5, value=entry['url'])
            ws.cell(row=row_idx, column=6, value=entry['reason'])

        for col in range(1, len(site_headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 28

    # Save
    scan_id = str(uuid.uuid4())[:8]
    USER_SCANNER_XLSX_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    filename = f'user_scanner_{ts}_{scan_id}.xlsx'
    filepath = USER_SCANNER_XLSX_DIR / filename
    wb.save(str(filepath))
    logger.info("Full scan XLSX saved: %s", filepath)
    return filepath


# ---------------------------------------------------------------------------
# Scan history – keep last N runs
# ---------------------------------------------------------------------------

def _load_history() -> List[Dict[str, Any]]:
    if not USER_SCANNER_HISTORY.exists():
        return []
    try:
        with USER_SCANNER_HISTORY.open('r', encoding='utf-8') as fh:
            data = json.load(fh)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_history(history: List[Dict[str, Any]]) -> None:
    try:
        USER_SCANNER_HISTORY.parent.mkdir(parents=True, exist_ok=True)
        with USER_SCANNER_HISTORY.open('w', encoding='utf-8') as fh:
            json.dump(history, fh, indent=2)
    except Exception as exc:
        logger.error('Failed to save scan history: %s', exc)


def record_scan_run(scan_result: Dict[str, Any], xlsx_path: Path) -> Dict[str, Any]:
    """Add a completed scan to the history and prune old entries."""
    scan_id = xlsx_path.stem
    entry = {
        'scanId': scan_id,
        'scannedAt': scan_result.get('scannedAt', ''),
        'totalEmployees': scan_result.get('totalEmployees', 0),
        'xlsxFile': xlsx_path.name,
        'registeredTotal': sum(
            r.get('registeredCount', 0) for r in scan_result.get('records', [])
        ),
    }

    history = _load_history()
    history.insert(0, entry)

    while len(history) > MAX_SCAN_HISTORY:
        removed = history.pop()
        old_file = USER_SCANNER_XLSX_DIR / removed.get('xlsxFile', '')
        if old_file.exists():
            try:
                old_file.unlink()
            except Exception:
                pass

    _save_history(history)
    return entry


def load_scan_history() -> List[Dict[str, Any]]:
    """Return the last N scan run metadata entries."""
    return _load_history()


def get_xlsx_path(scan_id: str) -> Optional[Path]:
    """Look up the XLSX file for a given scan ID."""
    history = _load_history()
    for entry in history:
        if entry.get('scanId') == scan_id:
            filepath = USER_SCANNER_XLSX_DIR / entry['xlsxFile']
            if filepath.exists():
                return filepath
    return None


def clear_scan_history() -> int:
    """Delete all scan history entries and their XLSX files.

    Returns the number of entries removed.
    """
    history = _load_history()
    count = len(history)
    for entry in history:
        xlsx_file = USER_SCANNER_XLSX_DIR / entry.get('xlsxFile', '')
        if xlsx_file.exists():
            try:
                xlsx_file.unlink()
            except Exception:
                pass
    _save_history([])
    logger.info('Cleared %d scan history entries', count)
    return count
