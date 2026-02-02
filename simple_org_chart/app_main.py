from flask import Flask, render_template, render_template_string, jsonify, request, send_from_directory, send_file, session, redirect, url_for
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_session import Session
import atexit
import json
import os
from datetime import datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover - Python < 3.9 fallback
    ZoneInfo = None  # type: ignore[assignment]
import threading
import time
from io import BytesIO
import logging
from dotenv import load_dotenv
from werkzeug.utils import secure_filename

import hashlib
import secrets
try:
    from PIL import Image
except ImportError:
    Image = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

import simple_org_chart.config as app_config
from simple_org_chart.auth import login_required, require_auth, sanitize_next_path
from simple_org_chart.settings import (
    DEFAULT_SETTINGS,
    department_is_ignored,
    employee_is_ignored,
    load_settings,
    normalize_filter_value,
    parse_ignored_departments,
    parse_ignored_employees,
    parse_ignored_titles,
    save_settings,
    translate_placeholder,
)
from simple_org_chart.msgraph import (
    calculate_days_since,
    collect_disabled_users,
    collect_last_login_records,
    datetime_to_iso,
    fetch_all_employees,
    fetch_employee_photo,
    get_access_token,
    parse_graph_datetime,
    _enrich_mailbox_metadata,
)
from simple_org_chart.reports import (
    ReportCacheManager,
    apply_disabled_filters,
    apply_filtered_user_filters,
    apply_last_login_filters,
    apply_missing_manager_filters,
    load_disabled_users_data,
    load_filtered_license_data,
    load_filtered_user_data,
    load_last_login_data,
    load_missing_manager_data,
    load_recently_hired_data,
)
from simple_org_chart.scheduler import (
    configure_scheduler,
    is_scheduler_running,
    restart_scheduler,
    start_scheduler,
    stop_scheduler,
)
from simple_org_chart.utils.files import validate_image_file

# Import from new split modules
from simple_org_chart.hierarchy import (
    build_org_hierarchy,
    collect_missing_manager_records,
    collect_employee_option_labels,
    collect_unique_field_values,
    flatten_hierarchy_to_employee_list,
    get_employee_list_for_metadata,
    load_cached_employees,
)
from simple_org_chart.data_update import (
    collect_recently_disabled_employees,
    collect_recently_hired_employees,
    load_data_update_status,
    mark_data_update_finished,
    mark_data_update_running,
    update_employee_data,
    _load_fetch_all_employees_fallback,
)
from simple_org_chart.exports import (
    build_microsip_directory_items,
    format_hire_date,
)

load_dotenv()

# Configure logging with ISO 8601 timestamp format
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s:%(name)s:%(message)s',
    datefmt='%Y-%m-%dT%H:%M:%S.000Z'
)
logger = logging.getLogger(__name__)

def _parse_port(raw_value, fallback):
    try:
        port = int(raw_value)
    except (TypeError, ValueError):
        if raw_value:
            logger.warning(f"Invalid APP_PORT value '{raw_value}', falling back to {fallback}")
        return fallback

    if not (1 <= port <= 65535):
        logger.warning(f"APP_PORT {port} out of range 1-65535, falling back to {fallback}")
        return fallback

    return port

DEFAULT_APP_PORT = 5000
APP_PORT = _parse_port(os.environ.get('APP_PORT'), DEFAULT_APP_PORT)
logger.info(f"Application port resolved to {APP_PORT}")

# Environment-configurable settings
SESSION_TYPE = os.environ.get('SESSION_TYPE', 'filesystem')
MAX_FILE_SIZE = int(os.environ.get('MAX_FILE_SIZE_MB', '5')) * 1024 * 1024
ALLOWED_EXTENSIONS = set(ext.strip().lower() for ext in os.environ.get('ALLOWED_LOGO_EXTENSIONS', 'png,jpg,jpeg').split(',') if ext.strip())
FAVICON_ALLOWED_EXTENSIONS = set(ext.strip().lower() for ext in os.environ.get('ALLOWED_FAVICON_EXTENSIONS', 'ico,png,jpg,jpeg').split(',') if ext.strip())
PHOTO_CACHE_SECONDS = int(os.environ.get('PHOTO_CACHE_SECONDS', '3600'))
PHOTO_CACHE_FILE_SECONDS = int(os.environ.get('PHOTO_CACHE_FILE_SECONDS', '86400'))
RATE_LIMIT_DEFAULT = os.environ.get('RATE_LIMIT_DEFAULT', '200 per day,50 per hour')
RATE_LIMIT_LOGIN = os.environ.get('RATE_LIMIT_LOGIN', '5 per minute')
RATE_LIMIT_PHOTO = os.environ.get('RATE_LIMIT_PHOTO', '500 per hour')
RATE_LIMIT_SETTINGS = os.environ.get('RATE_LIMIT_SETTINGS', '20 per minute')
RATE_LIMIT_UPLOAD = os.environ.get('RATE_LIMIT_UPLOAD', '5 per minute')
RATE_LIMIT_REFRESH = os.environ.get('RATE_LIMIT_REFRESH', '1 per minute')

# Security headers (configurable via .env)
SECURITY_HEADER_CONTENT_TYPE_OPTIONS = os.environ.get('SECURITY_HEADER_CONTENT_TYPE_OPTIONS', 'nosniff')
SECURITY_HEADER_FRAME_OPTIONS = os.environ.get('SECURITY_HEADER_FRAME_OPTIONS', 'DENY')
SECURITY_HEADER_XSS_PROTECTION = os.environ.get('SECURITY_HEADER_XSS_PROTECTION', '1; mode=block')
SECURITY_HEADER_HSTS = os.environ.get('SECURITY_HEADER_HSTS', 'max-age=31536000; includeSubDomains')
SECURITY_HEADER_CSP = os.environ.get('SECURITY_HEADER_CSP', "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; img-src 'self' data:;")

app = Flask(
    __name__,
    static_folder=str(app_config.STATIC_DIR),
    template_folder=str(app_config.TEMPLATE_DIR),
)

_allowed_origins = [origin.strip() for origin in os.environ.get('CORS_ALLOWED_ORIGINS', '').split(',') if origin.strip()]
if _allowed_origins:
    CORS(app, resources={r"/api/*": {"origins": _allowed_origins}})

# Security Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_TYPE'] = SESSION_TYPE
app.config['SESSION_PERMANENT'] = False

# Initialize extensions
Session(app)


def _is_localhost_exempt():
    """Check if request should be exempt from rate limiting (localhost/health checks)."""
    remote_addr = get_remote_address()
    return remote_addr in ('127.0.0.1', '::1', 'localhost')


limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[limit.strip() for limit in RATE_LIMIT_DEFAULT.split(',') if limit.strip()],
    storage_uri="memory://"
)

# Exempt localhost from all rate limits (health checks, internal requests)
limiter.exempt(_is_localhost_exempt)

# Simple authentication settings
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD')
if not ADMIN_PASSWORD:
    raise RuntimeError('ADMIN_PASSWORD environment variable must be set to a strong value')
if ADMIN_PASSWORD in {'admin123', 'your-admin-password-here'}:
    raise RuntimeError('ADMIN_PASSWORD must not use the default placeholder value')

# Security headers
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = SECURITY_HEADER_CONTENT_TYPE_OPTIONS
    response.headers['X-Frame-Options'] = SECURITY_HEADER_FRAME_OPTIONS
    response.headers['X-XSS-Protection'] = SECURITY_HEADER_XSS_PROTECTION
    response.headers['Strict-Transport-Security'] = SECURITY_HEADER_HSTS
    response.headers['Content-Security-Policy'] = SECURITY_HEADER_CSP
    return response
app_config.ensure_directories()

DATA_DIR = str(app_config.DATA_DIR)
SETTINGS_FILE = str(app_config.SETTINGS_FILE)
DATA_FILE = str(app_config.DATA_FILE)
MISSING_MANAGER_FILE = str(app_config.MISSING_MANAGER_FILE)
EMPLOYEE_LIST_FILE = str(app_config.EMPLOYEE_LIST_FILE)
DISABLED_LICENSE_FILE = str(app_config.DISABLED_LICENSE_FILE)
FILTERED_LICENSE_FILE = str(app_config.FILTERED_LICENSE_FILE)
FILTERED_USERS_FILE = str(app_config.FILTERED_USERS_FILE)
DISABLED_USERS_FILE = str(app_config.DISABLED_USERS_FILE)
LAST_LOGIN_FILE = str(app_config.LAST_LOGIN_FILE)
RECENTLY_DISABLED_FILE = str(app_config.RECENTLY_DISABLED_FILE)
RECENTLY_HIRED_FILE = str(app_config.RECENTLY_HIRED_FILE)
DATA_UPDATE_STATUS_FILE = os.path.join(DATA_DIR, 'data_update_status.json')

logger.info(f"DATA_DIR set to: {DATA_DIR}")

TENANT_ID = os.environ.get('AZURE_TENANT_ID')
CLIENT_ID = os.environ.get('AZURE_CLIENT_ID')
CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET')

if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
    logger.warning("Missing Azure AD credentials in environment variables!")
    logger.warning("AZURE_TENANT_ID: " + ("Set" if TENANT_ID else "Not set"))
    logger.warning("AZURE_CLIENT_ID: " + ("Set" if CLIENT_ID else "Not set"))
    logger.warning("AZURE_CLIENT_SECRET: " + ("Set" if CLIENT_SECRET else "Not set"))
    logger.warning("Please check your .env file exists and contains the correct values")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


configure_scheduler(update_employee_data)
report_cache = ReportCacheManager(refresh_callback=update_employee_data)


if hasattr(app, 'before_serving'):

    @app.before_serving
    def _start_scheduler_when_ready():
        start_scheduler()


    @app.after_serving
    def _stop_scheduler_on_shutdown():
        stop_scheduler()

elif hasattr(app, 'before_request'):

    @app.before_request
    def _ensure_scheduler_started():
        if not is_scheduler_running():
            start_scheduler()


atexit.register(stop_scheduler)

def get_template(template_name):
    """Load HTML template from file"""
    possible_paths = [
        f'templates/{template_name}',
        template_name,
        os.path.join(os.path.dirname(__file__), 'templates', template_name),
        os.path.join(os.path.dirname(__file__), template_name),
        os.path.join(str(app_config.TEMPLATE_DIR), template_name)
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    # Only log template loading in debug mode to reduce log spam
                    logger.debug(f"Loading template from: {path}")
                    return f.read()
            except Exception as e:
                logger.error(f"Error reading {path}: {e}")
    
    logger.error(f"{template_name} not found in any expected location")
    return f"<h1>Error: {template_name} not found</h1>"

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit(RATE_LIMIT_LOGIN)
def login():
    next_page = sanitize_next_path(request.args.get('next', ''))

    if request.method == 'POST':
        try:
            payload = request.get_json(silent=True)
            password = None

            if isinstance(payload, dict):
                password = payload.get('password')

            if password is None:
                password = request.form.get('password')

            password = password or ''

            logger.info(f"Login attempt from {get_remote_address()} - IP address authentication check")

            if password == ADMIN_PASSWORD:
                session['authenticated'] = True
                session['username'] = 'admin'
                logger.info("Successful login")
                return jsonify({
                    'success': True,
                    'next': next_page or ''
                })

            logger.warning("Failed login attempt - password mismatch")
            return jsonify({'error': 'Invalid password'}), 401
        except Exception as e:
            logger.error(f"Login error: {e}")
            return jsonify({'error': 'Login failed'}), 500

    settings = load_settings()
    chart_title = (settings.get('chartTitle') or '').strip() or 'DB AutoOrgChart'
    favicon_path = settings.get('faviconPath', '/favicon.ico')

    return render_template(
        'login.html',
        chart_title=chart_title,
        next_page=next_page,
        favicon_path=favicon_path
    )

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/')
def index():
    template_content = get_template('index.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')
    
    # Inject favicon link into template
    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')
    
    return render_template_string(template_content)

@app.route('/configure')
def configure():
    if not session.get('authenticated'):
        # Redirect to login preserving the intended destination
        desired_path = sanitize_next_path(request.path)
        params = {'next': desired_path} if desired_path else {}
        return redirect(url_for('login', **params))
    
    template_content = get_template('configure.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')
    logo_path = settings.get('logoPath', '/static/icon.png')
    chart_title = (settings.get('chartTitle') or '').strip() or 'DB AutoOrgChart'
    
    # Inject favicon link into template
    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')
    
    return render_template_string(
        template_content,
        chart_title=chart_title,
        logo_path=logo_path,
        favicon_image_path=favicon_path,
        _=translate_placeholder
    )


@app.route('/reports')
@login_required
def reports():
    template_content = get_template('reports.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')

    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')

    return render_template_string(template_content)

@app.route('/static/icon_custom_<string:file_hash>.png')
def serve_custom_logo(file_hash):
    """Serve custom logo files from the data directory"""
    # Validate file_hash to prevent directory traversal
    if not file_hash.isalnum() or len(file_hash) != 8:
        return "Invalid logo identifier", 400
    
    custom_logo = os.path.join(DATA_DIR, f'icon_custom_{file_hash}.png')
    
    if os.path.exists(custom_logo) and os.path.isfile(custom_logo):
        return send_file(custom_logo, 
                       mimetype='image/png',
                       as_attachment=False,
                       max_age=PHOTO_CACHE_SECONDS)
    else:
        return "Logo not found", 404

@app.route('/static/favicon_custom_<file_hash>.<ext>')
def serve_custom_favicon(file_hash, ext):
    """Serve custom favicon files from the data directory"""
    # Validate file_hash to prevent directory traversal
    if not file_hash.isalnum() or len(file_hash) != 12:
        return "Invalid favicon identifier", 400
    
    # Validate extension
    if ext not in ['ico', 'png']:
        return "Invalid favicon format", 400
    
    custom_favicon = os.path.join(DATA_DIR, f'favicon_custom_{file_hash}.{ext}')
    
    if os.path.exists(custom_favicon) and os.path.isfile(custom_favicon):
        mimetype = 'image/x-icon' if ext == 'ico' else 'image/png'
        return send_file(custom_favicon, 
                       mimetype=mimetype,
                       as_attachment=False,
                       max_age=PHOTO_CACHE_SECONDS)
    else:
        return "Favicon not found", 404

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve regular static files"""
    return send_from_directory(app.static_folder, filename)

@app.route('/api/photo/<user_id>')
@limiter.limit(RATE_LIMIT_PHOTO)
def get_employee_photo(user_id):
    """Serve employee photo from Microsoft Graph API with caching"""
    try:
        # Create photos cache directory
        photos_dir = os.path.join(DATA_DIR, 'photos')
        if not os.path.exists(photos_dir):
            os.makedirs(photos_dir, exist_ok=True)
        
        # Check if photo is cached
        photo_file = os.path.join(photos_dir, f"{user_id}.jpg")
        
        if os.path.exists(photo_file):
            # Check if cache is still valid
            if time.time() - os.path.getmtime(photo_file) < PHOTO_CACHE_FILE_SECONDS:
                response = send_file(photo_file, mimetype='image/jpeg')
                # Add cache headers to prevent browser caching issues
                response.headers['Cache-Control'] = f'public, max-age={PHOTO_CACHE_SECONDS}'
                response.headers['Last-Modified'] = datetime.fromtimestamp(os.path.getmtime(photo_file)).strftime('%a, %d %b %Y %H:%M:%S GMT')
                return response
        
        # Fetch fresh photo from Graph API
        token = get_access_token()
        if token:
            photo_data = fetch_employee_photo(user_id, token)
            if photo_data:
                # Save to cache
                with open(photo_file, 'wb') as f:
                    f.write(photo_data)
                
                # Serve the photo
                response = send_file(
                    BytesIO(photo_data),
                    mimetype='image/jpeg',
                    as_attachment=False
                )
                response.headers['Cache-Control'] = f'public, max-age={PHOTO_CACHE_SECONDS}'
                return response
        
        # Fallback to default user icon
        logger.debug(f"No photo available for user {user_id}, using fallback")
        return send_from_directory(app.static_folder, 'usericon.png')
        
    except Exception as e:
        logger.error(f"Error serving photo for user {user_id}: {e}")
        return send_from_directory(app.static_folder, 'usericon.png')

@app.route('/api/employees')
def get_employees():
    try:
        logger.info("API request for /api/employees received")
        if not os.path.exists(DATA_FILE):
            logger.info("Data file does not exist, attempting to create it...")
            update_employee_data(source='api-employees')
            
        # Double check the file exists after update attempt
        if not os.path.exists(DATA_FILE):
            logger.error(f"Could not create data file {DATA_FILE}")
            return jsonify({'error': 'No employee data available. Please check configuration.'}), 500
        
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)

        settings = load_settings()
        months_threshold = settings.get('newEmployeeMonths', 3)

        session_override_present = 'topUserEmail' in session
        session_top_user = (session.get('topUserEmail') or '').strip() if session_override_present else None
        settings_top_user_email = (settings.get('topLevelUserEmail') or '').strip()

        requested_top_user = None
        override_reason = None

        if session_override_present:
            requested_top_user = session_top_user
            override_reason = 'session override'
        elif settings_top_user_email:
            current_root_email = (data or {}).get('email') or ''
            if current_root_email.strip().lower() != settings_top_user_email.lower():
                requested_top_user = settings_top_user_email
                override_reason = 'settings default enforcement'

        if requested_top_user is not None:
            employees = load_cached_employees()
            if not employees and data:
                employees = flatten_hierarchy_to_employee_list(data)
            if not employees:
                logger.info("Employee cache unavailable; fetching employees from Graph API for top user override")
                employees, _, _ = fetch_all_employees(
                    fallback_loader=_load_fetch_all_employees_fallback,
                )
                if employees:
                    try:
                        with open(EMPLOYEE_LIST_FILE, 'w') as cache_file:
                            json.dump(employees, cache_file, indent=2)
                    except Exception as cache_error:
                        logger.error(f"Failed to refresh employee cache: {cache_error}")

            if employees:
                override_hierarchy = build_org_hierarchy(
                    employees,
                    top_user_email_override=requested_top_user,
                    settings=settings
                )
                if override_hierarchy:
                    data = override_hierarchy

                    if override_reason == 'settings default enforcement':
                        try:
                            with open(DATA_FILE, 'w') as data_file:
                                json.dump(data, data_file, indent=2)
                            missing_records = collect_missing_manager_records(
                                employees,
                                data,
                                settings,
                                top_user_email_override=requested_top_user
                            )
                            with open(MISSING_MANAGER_FILE, 'w') as report_file:
                                json.dump(missing_records, report_file, indent=2)
                            logger.info("Refreshed global hierarchy cache to align with environment top user")
                        except Exception as cache_error:
                            logger.error(f"Failed to persist environment-aligned hierarchy: {cache_error}")
                else:
                    logger.warning("Failed to build hierarchy with requested top user override; returning cached hierarchy")
            else:
                logger.warning("Unable to locate employee data while applying top user override; returning cached hierarchy")
        
        if data:
            def update_new_status(node):
                if node.get('hireDate'):
                    try:
                        hire_date = datetime.fromisoformat(node['hireDate'])
                        if hire_date.tzinfo:
                            cutoff_date = datetime.now(hire_date.tzinfo) - timedelta(days=months_threshold * 30)
                        else:
                            cutoff_date = datetime.now() - timedelta(days=months_threshold * 30)
                        node['isNewEmployee'] = hire_date > cutoff_date
                    except:
                        node['isNewEmployee'] = False
                else:
                    node['isNewEmployee'] = False
                
                if node.get('children'):
                    for child in node['children']:
                        update_new_status(child)
            
            update_new_status(data)
        
        # Debug logging for root user
        if data and data.get('name'):
            logger.info(f"Returning org chart data with root user: {data['name']} ({data.get('email', 'no email')})")
        
        if not data:
            logger.warning("No hierarchical data available")
            employees, _, _ = fetch_all_employees(
                fallback_loader=_load_fetch_all_employees_fallback,
            )
            if employees:
                data = {
                    'id': 'root',
                    'name': 'Organization',
                    'title': 'All Employees',
                    'department': '',
                    'email': '',
                    'phone': '',
                    'businessPhone': '',
                    'location': '',
                    'officeLocation': '',
                    'city': '',
                    'state': '',
                    'country': '',
                    'fullAddress': '',
                    'children': employees
                }
            else:
                data = {
                    'id': 'root',
                    'name': 'No Data',
                    'title': 'Please check configuration',
                    'businessPhone': '',
                    'children': []
                }
        
        return jsonify(data)
    except Exception as e:
        logger.error(f"Error in get_employees: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/<filename>.json')
def get_microsip_directory(filename):
    """Expose a MicroSIP-compatible directory derived from cached employees."""
    settings = load_settings()
    
    # Check if the requested filename matches the configured one
    configured_filename = (settings.get('directoryJsonFilename') or 'contacts').strip() or 'contacts'
    if filename != configured_filename:
        return jsonify({'error': 'Not found'}), 404
    
    # Check if the feed is enabled
    if not settings.get('directoryJsonEnabled', False):
        logger.info(f"MicroSIP directory request denied - feed is disabled")
        return jsonify({'error': 'Directory feed is disabled'}), 403
    
    logger.info("Request received for MicroSIP directory export")

    employees = get_employee_list_for_metadata()
    if not employees:
        logger.warning("No cached employees available for MicroSIP directory; attempting refresh")
        try:
            update_employee_data(source='contacts-refresh')
        except Exception as refresh_error:
            logger.error(f"Failed to refresh employee cache for MicroSIP directory: {refresh_error}")
        employees = get_employee_list_for_metadata()

    items = build_microsip_directory_items(employees, settings=settings)
    payload = {
        'refresh': int(time.time()),
        'items': items,
    }

    response = jsonify(payload)
    response.headers['Cache-Control'] = 'no-store'

    if not items:
        logger.warning("MicroSIP directory generated without any entries")

    return response


@app.route('/<filename>.xml')
def get_contacts_xml(filename):
    """Expose a phone-compatible remote phonebook XML derived from cached employees.
    
    Query parameters:
        format: Phone vendor format (default: yealink). Currently supported: yealink
    
    Yealink T31P setup:
        1. Open phone web UI -> Directory -> Remote Phone Book
        2. Set Remote URL to: http://<server>/<filename>.xml
        3. Set Display Name to: Company Directory
        4. Save and reboot if prompted
    """
    from simple_org_chart.exports import build_yealink_phonebook_xml

    settings = load_settings()
    
    # Check if the requested filename matches the configured one
    configured_filename = (settings.get('directoryXmlFilename') or 'contacts').strip() or 'contacts'
    if filename != configured_filename:
        return app.response_class(
            response='<?xml version="1.0" encoding="UTF-8"?>\n<Error>Not found</Error>',
            status=404,
            mimetype='application/xml'
        )
    
    # Check if the feed is enabled
    if not settings.get('directoryXmlEnabled', False):
        logger.info(f"XML directory request denied - feed is disabled")
        return app.response_class(
            response='<?xml version="1.0" encoding="UTF-8"?>\n<Error>Directory feed is disabled</Error>',
            status=403,
            mimetype='application/xml'
        )

    phone_format = request.args.get('format', 'yealink').lower()

    if phone_format != 'yealink':
        logger.warning(f"Unsupported contacts.xml format requested: {phone_format}")
        return app.response_class(
            response='<?xml version="1.0" encoding="UTF-8"?>\n<Error>Unsupported format. Supported: yealink</Error>',
            status=400,
            mimetype='application/xml'
        )

    logger.info(f"Request received for contacts.xml (format={phone_format})")

    employees = get_employee_list_for_metadata()
    if not employees:
        logger.warning("No cached employees available for XML directory; attempting refresh")
        try:
            update_employee_data(source='contacts-xml-refresh')
        except Exception as refresh_error:
            logger.error(f"Failed to refresh employee cache for XML directory: {refresh_error}")
        employees = get_employee_list_for_metadata()

    settings = load_settings()
    chart_title = settings.get('chartTitle', 'Organization Directory')
    xml_content = build_yealink_phonebook_xml(employees, settings=settings, title=chart_title)

    response = app.response_class(
        response=xml_content,
        status=200,
        mimetype='application/xml'
    )
    response.headers['Cache-Control'] = 'no-store'
    response.headers['Content-Disposition'] = 'inline; filename="contacts.xml"'

    if not employees:
        logger.warning("XML phonebook generated without any entries")

    return response


@app.route('/api/settings', methods=['GET', 'POST'])
def handle_settings():
    if request.method == 'GET':
        # GET is allowed without auth for loading initial settings
        settings = load_settings()
        if 'topUserEmail' in session:
            settings['topUserEmail'] = session.get('topUserEmail') or ''
        data_last_updated = None
        try:
            if os.path.exists(DATA_FILE):
                modified_at = datetime.fromtimestamp(os.path.getmtime(DATA_FILE), tz=timezone.utc)
                data_last_updated = modified_at.isoformat()
        except Exception as timestamp_error:
            logger.warning("Failed to compute data last updated timestamp: %s", timestamp_error)

        settings['dataLastUpdatedAt'] = data_last_updated
        settings['dataUpdateStatus'] = load_data_update_status()

        return jsonify(settings)
    
    elif request.method == 'POST':
        # POST requires authentication
        if not session.get('authenticated'):
            return jsonify({'error': 'Authentication required'}), 401
            
        try:
            # Simply update settings without validation
            new_settings = request.json
            current_settings = load_settings()
            current_settings.update(new_settings)
            
            if save_settings(current_settings):
                if ('updateTime' in new_settings or 'autoUpdateEnabled' in new_settings):
                    threading.Thread(target=restart_scheduler).start()
                
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'Failed to save settings'}), 500
        except Exception as e:
            logger.error(f"Error updating settings: {e}")
            return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/metadata/options')
@require_auth
def get_metadata_options():
    employees = get_employee_list_for_metadata()
    job_titles = collect_unique_field_values(employees, 'title')
    departments = collect_unique_field_values(employees, 'department')
    employee_options = collect_employee_option_labels(employees)

    return jsonify({
        'jobTitles': job_titles,
        'departments': departments,
        'employees': employee_options
    })

@app.route('/api/set-top-user', methods=['POST'])
@limiter.limit(RATE_LIMIT_SETTINGS)
def set_top_user():
    """Store the caller's preferred top-level user in their session"""
    try:
        data = request.json or {}
        if 'topUserEmail' not in data:
            return jsonify({'error': 'Missing topUserEmail parameter'}), 400

        requested_email = (data.get('topUserEmail') or '').strip()
        session['topUserEmail'] = requested_email
        session.modified = True

        logger.info(
            f"Stored session-specific top user preference '{requested_email or 'auto-detect'}' for client {request.remote_addr}"
        )

        return jsonify({'success': True, 'topUserEmail': requested_email})
    except Exception as e:
        logger.error(f"Error updating top-level user session preference: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/set-multiline-enabled', methods=['POST'])
@require_auth
@limiter.limit(RATE_LIMIT_SETTINGS)
def set_multiline_enabled():
    """Public endpoint to toggle multi-line children layout (Compact Teams)."""
    try:
        data = request.json or {}
        if 'multiLineChildrenEnabled' not in data:
            return jsonify({'error': 'Missing multiLineChildrenEnabled parameter'}), 400

        current_settings = load_settings()
        current_settings['multiLineChildrenEnabled'] = bool(data['multiLineChildrenEnabled'])

        if save_settings(current_settings):
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Failed to save settings'}), 500
    except Exception as e:
        logger.error(f"Error updating multi-line setting: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/test-hierarchy/<email>')
@require_auth
@limiter.limit(RATE_LIMIT_UPLOAD)
def test_hierarchy(email):
    """Test endpoint to check hierarchy building with specific email"""
    try:
        # Temporarily override settings for testing
        import tempfile
        
        # Fetch fresh employees
        employees, _, _ = fetch_all_employees(
            fallback_loader=_load_fetch_all_employees_fallback,
        )
        if not employees:
            return jsonify({'error': 'No employees found'}), 404
            
        # Find the user with this email
        target_user = None
        for emp in employees:
            if emp.get('email') == email:
                target_user = emp
                break
                
        if not target_user:
            return jsonify({'error': f'User with email {email} not found'}), 404
            
        # Build hierarchy with specified top-level user
        hierarchy = build_org_hierarchy(employees, top_user_email_override=email)
        
        if hierarchy:
            return jsonify({
                'success': True,
                'root_user': {
                    'name': hierarchy.get('name'),
                    'email': hierarchy.get('email'),
                    'title': hierarchy.get('title')
                },
                'test_email': email,
                'target_user': {
                    'name': target_user.get('name'),
                    'email': target_user.get('email'),
                    'title': target_user.get('title')
                },
                'total_employees': len(employees)
            })
        else:
            return jsonify({'error': 'Failed to build hierarchy'}), 500
            
    except Exception as e:
        logger.error(f"Error in test hierarchy: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-logo', methods=['POST'])
@require_auth
@limiter.limit(RATE_LIMIT_UPLOAD)
def upload_logo():
    try:
        if 'logo' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['logo']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': f'File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB'}), 400
        
        if file and allowed_file(file.filename):
            # Validate it's actually a safe image
            if not validate_image_file(file):
                return jsonify({'error': 'Invalid or corrupted image file'}), 400
            
            # Generate secure filename
            filename = secure_filename(file.filename)
            file_hash = hashlib.md5(file.read()).hexdigest()[:8]
            file.seek(0)  # Reset after reading for hash
            
            custom_logo_path = os.path.join(DATA_DIR, f'icon_custom_{file_hash}.png')
            
            # Check if directory is writable
            if not os.access(DATA_DIR, os.W_OK):
                return jsonify({'error': 'Server configuration error'}), 500

            file.save(custom_logo_path)
            
            # Verify file was saved
            if not os.path.exists(custom_logo_path):
                return jsonify({'error': 'Failed to save file'}), 500
            
            settings = load_settings()
            settings['logoPath'] = f'/static/icon_custom_{file_hash}.png'
            save_settings(settings)
            
            return jsonify({'success': True, 'path': settings['logoPath']})
        else:
            return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG allowed'}), 400
    except Exception as e:
        logger.error(f"Error uploading logo: {e}")
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/reset-logo', methods=['POST'])
@require_auth
def reset_logo():
    try:
        # Remove any custom logo files from data directory
        import glob
        custom_logos = glob.glob(os.path.join(DATA_DIR, 'icon_custom*.png'))
        for logo_path in custom_logos:
            if os.path.exists(logo_path):
                os.remove(logo_path)
        
        settings = load_settings()
        settings['logoPath'] = '/static/icon.png'
        save_settings(settings)
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting logo: {e}")
        return jsonify({'error': 'Reset failed'}), 500

@app.route('/api/upload-favicon', methods=['POST'])
@require_auth
@limiter.limit(RATE_LIMIT_UPLOAD)
def upload_favicon():
    try:
        if 'favicon' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['favicon']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': f'File size exceeds {MAX_FILE_SIZE // (1024*1024)}MB limit'}), 400
        
        # Check file extension
        filename = secure_filename(file.filename)
        file_ext = filename.lower().split('.')[-1] if '.' in filename else ''
        
        if file_ext in FAVICON_ALLOWED_EXTENSIONS or file.content_type in ['image/x-icon', 'image/png', 'image/jpeg']:
            # Create unique filename
            import hashlib
            file_content = file.read()
            file_hash = hashlib.md5(file_content).hexdigest()[:12]
            
            # Determine file extension based on content type or filename
            if file.content_type == 'image/x-icon' or file_ext == 'ico':
                ext = 'ico'
            elif file.content_type == 'image/png' or file_ext == 'png':
                ext = 'png'
            else:
                ext = 'png'  # Default to PNG for jpg/jpeg
            
            favicon_filename = f'favicon_custom_{file_hash}.{ext}'
            favicon_path = os.path.join(DATA_DIR, favicon_filename)
            
            # Validate and process image
            try:
                from PIL import Image
                import io
                
                # Open and validate image
                image = Image.open(io.BytesIO(file_content))
                
                # Convert to appropriate format and resize if needed
                if ext == 'ico':
                    # For ICO files, keep original if it's already ICO, otherwise convert
                    if image.format != 'ICO':
                        # Resize to 32x32 for favicon
                        image = image.resize((32, 32), Image.Resampling.LANCZOS)
                        # Convert to RGBA if needed
                        if image.mode != 'RGBA':
                            image = image.convert('RGBA')
                        image.save(favicon_path, 'ICO', sizes=[(32, 32)])
                    else:
                        with open(favicon_path, 'wb') as f:
                            f.write(file_content)
                else:
                    # For PNG files
                    if image.size != (32, 32):
                        image = image.resize((32, 32), Image.Resampling.LANCZOS)
                    if image.mode != 'RGBA':
                        image = image.convert('RGBA')
                    image.save(favicon_path, 'PNG', optimize=True)
                
            except Exception as img_error:
                logger.error(f"Image processing error: {img_error}")
                return jsonify({'error': 'Invalid image file'}), 400
            
            if os.path.exists(favicon_path):
                settings = load_settings()
                settings['faviconPath'] = f'/static/favicon_custom_{file_hash}.{ext}'
                save_settings(settings)
                
                return jsonify({'success': True, 'path': settings['faviconPath']})
            else:
                return jsonify({'error': 'Failed to save file'}), 500
        else:
            return jsonify({'error': 'Invalid file type. Only ICO, PNG, JPG, JPEG allowed'}), 400
    except Exception as e:
        logger.error(f"Error uploading favicon: {e}")
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/reset-favicon', methods=['POST'])
@require_auth
def reset_favicon():
    try:
        # Remove any custom favicon files from data directory
        import glob
        custom_favicons = glob.glob(os.path.join(DATA_DIR, 'favicon_custom*'))
        for favicon_path in custom_favicons:
            if os.path.exists(favicon_path):
                os.remove(favicon_path)
        
        settings = load_settings()
        settings['faviconPath'] = '/favicon.ico'
        save_settings(settings)
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting favicon: {e}")
        return jsonify({'error': 'Reset failed'}), 500

@app.route('/api/reset-all-settings', methods=['POST'])
@require_auth
def reset_all_settings():
    try:
        # Remove any custom logo files
        import glob
        custom_logos = glob.glob(os.path.join(DATA_DIR, 'icon_custom*.png'))
        for logo_path in custom_logos:
            if os.path.exists(logo_path):
                os.remove(logo_path)
        
        save_settings(DEFAULT_SETTINGS)
        
        threading.Thread(target=restart_scheduler).start()
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting all settings: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-xlsx')
def export_xlsx():
    """Export organizational data to XLSX format"""
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500
    
    try:
        # Load employee data
        if not os.path.exists(DATA_FILE):
            update_employee_data(source='export-xlsx')
        
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        if not data:
            return jsonify({'error': 'No employee data available'}), 404
        
        # Load settings to check filtering preferences and column visibility
        settings = load_settings()
        hide_disabled_users = settings.get('hideDisabledUsers', True)
        hide_guest_users = settings.get('hideGuestUsers', True)
        hide_no_title = settings.get('hideNoTitle', True)
        ignored_departments = parse_ignored_departments(settings)
        export_column_settings = settings.get('exportXlsxColumns', {}) or {}
        is_admin = bool(session.get('authenticated'))

        column_definitions = [
            ('name', 'Name', lambda node, manager: node.get('name', '')),
            ('title', 'Title', lambda node, manager: node.get('title', '')),
            ('department', 'Department', lambda node, manager: node.get('department', '')),
            ('email', 'Email', lambda node, manager: node.get('email', '')),
            ('phone', 'Phone', lambda node, manager: node.get('phone', '')),
            ('businessPhone', 'Business Phone', lambda node, manager: node.get('businessPhone', '')),
            ('hireDate', 'Hire Date', lambda node, manager: format_hire_date(node.get('hireDate', ''))),
            ('country', 'Country', lambda node, manager: node.get('country', '')),
            ('state', 'State', lambda node, manager: node.get('state', '')),
            ('city', 'City', lambda node, manager: node.get('city', '')),
            ('office', 'Office', lambda node, manager: node.get('officeLocation', '')),
            ('manager', 'Manager', lambda node, manager: manager)
        ]

        def column_is_visible(key):
            raw_mode = export_column_settings.get(key, 'show')
            mode = str(raw_mode).lower()
            normalized_admin = mode.replace('_', '').replace('-', '')
            if mode == 'hide':
                return False
            if mode == 'admin' and not is_admin:
                return False
            if normalized_admin in {'showadminonly', 'adminonly'} and not is_admin:
                return False
            return True

        visible_columns = [col for col in column_definitions if column_is_visible(col[0])]
        if not visible_columns:
            # Always include at least the Name column to avoid empty exports
            visible_columns = [column_definitions[0]]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Organization Chart"

        # Add headers with styling
        for col_index, (_, header, _) in enumerate(visible_columns, 1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Function to flatten organizational structure
        def flatten_org_data(node, manager_name="", row_num=2):
            if not node:
                return row_num
            
            # Check if we should skip this employee based on filtering settings
            title = node.get('title', '')
            department = node.get('department', '')
            account_enabled = node.get('accountEnabled', True)
            user_type = (node.get('userType') or '').lower()
            should_skip = (hide_no_title and (not title or title.strip() == '' or title.strip() == 'No Title')) or \
                         (department_is_ignored(department, ignored_departments)) or \
                         (hide_disabled_users and not account_enabled) or \
                         (hide_guest_users and user_type == 'guest')
            
            # Add current employee only if not filtering them out
            if not should_skip:
                for col_index, (_, _, extractor) in enumerate(visible_columns, 1):
                    ws.cell(row=row_num, column=col_index, value=extractor(node, manager_name))
                row_num += 1
            
            # Add children (using current employee name as manager if not skipped, otherwise pass through current manager)
            current_manager = node.get('name', '') if not should_skip else manager_name
            for child in node.get('children', []):
                row_num = flatten_org_data(child, current_manager, row_num)
                
            return row_num
        
        # Flatten the data starting from root
        flatten_org_data(data)
        
        # Auto-adjust column widths
        for col in range(1, len(visible_columns) + 1):
            column = get_column_letter(col)
            ws.column_dimensions[column].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"org-chart-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error exporting to XLSX: {e}")
        return jsonify({'error': 'Failed to export XLSX'}), 500
def _get_disabled_records_from_request(*, force_refresh=False, apply_filters=True):
    licensed_only = request.args.get('licensedOnly', 'true').lower() == 'true'
    include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
    include_members = request.args.get('includeMembers', 'true').lower() == 'true'
    recent_days_raw = request.args.get('recentDays')
    recent_days = None
    if recent_days_raw not in (None, ''):
        try:
            recent_days = int(recent_days_raw)
        except ValueError:
            logger.warning(f"Invalid recentDays value provided: {recent_days_raw}")

    records = load_disabled_users_data(report_cache, force_refresh=force_refresh)
    filtered_records = (
        apply_disabled_filters(
            records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests,
            include_members=include_members
        )
        if apply_filters else records
    )

    filter_payload = {
        'licensedOnly': licensed_only,
        'recentDays': recent_days,
        'includeGuests': include_guests,
        'includeMembers': include_members
    }

    return filtered_records, filter_payload


@app.route('/api/reports/missing-manager')
@require_auth
def get_missing_manager_report():
    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=False)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=False)
        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=False)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=False)

        records = load_missing_manager_data(report_cache, force_refresh=refresh)
        filtered_records = apply_missing_manager_filters(
            records,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
        )
        generated_at = None
        if os.path.exists(MISSING_MANAGER_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(MISSING_MANAGER_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'includeUserMailboxes': include_user_mailboxes,
                'includeSharedMailboxes': include_shared_mailboxes,
                'includeRoomEquipmentMailboxes': include_room_equipment_mailboxes,
                'includeEnabled': include_enabled,
                'includeDisabled': include_disabled,
                'includeLicensed': include_licensed,
                'includeUnlicensed': include_unlicensed,
                'includeMembers': include_members,
                'includeGuests': include_guests,
            }
        })
    except Exception as e:
        logger.error(f"Error loading missing manager report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/missing-manager/export')
@require_auth
def export_missing_manager_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=False)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=False)
        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=False)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=False)

        records = load_missing_manager_data(report_cache, force_refresh=refresh)
        filtered_records = apply_missing_manager_filters(
            records,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Managers"

        headers = [
            ('name', 'Name'),
            ('title', 'Title'),
            ('department', 'Department'),
            ('email', 'Email'),
            ('businessPhone', 'Business Phone'),
            ('location', 'Location'),
            ('managerName', 'Manager Name'),
            ('reason', 'Reason')
        ]

        reason_labels = {
            'no_manager': 'No manager assigned',
            'manager_not_found': 'Manager not found in data',
            'detached': 'Detached from hierarchy',
            'filtered': 'Filtered'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(filtered_records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'reason':
                    value = reason_labels.get(value, value or '')
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"missing-managers-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting missing manager report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-users')
@require_auth
def get_disabled_users_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        filtered_records, applied_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=True
        )

        generated_at = None
        if os.path.exists(DISABLED_USERS_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(DISABLED_USERS_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': applied_filters
        })
    except Exception as e:
        logger.error(f"Error loading disabled users report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-users/export')
@require_auth
def export_disabled_users_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records, _ = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=True
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('disabledDate', 'First Observed Disabled'),
            ('disabledDays', 'Days Since Observed Disabled'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'disabledDate' and value:
                    dt = parse_graph_datetime(value)
                    value = dt.date().isoformat() if dt else value
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting disabled users report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-this-year')
@require_auth
def get_recently_disabled_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        recent_days_raw = request.args.get('recentDays')
        if recent_days_raw in (None, ''):
            recent_days = 365
        else:
            try:
                recent_days = int(recent_days_raw)
            except ValueError:
                logger.warning(f"Invalid recentDays value provided for recently disabled report: {recent_days_raw}")
                recent_days = 365

        licensed_only = request.args.get('licensedOnly', 'false').lower() == 'true'
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'

        records = apply_disabled_filters(
            all_records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests
        )
        generated_at = None
        if os.path.exists(RECENTLY_DISABLED_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(RECENTLY_DISABLED_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': licensed_only,
                'recentDays': recent_days,
                'includeGuests': include_guests
            }
        })
    except Exception as e:
        logger.error(f"Error loading recently disabled report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-this-year/export')
@require_auth
def export_recently_disabled_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, _ = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        recent_days_raw = request.args.get('recentDays')
        if recent_days_raw in (None, ''):
            recent_days = 365
        else:
            try:
                recent_days = int(recent_days_raw)
            except ValueError:
                logger.warning(f"Invalid recentDays value provided for recently disabled export: {recent_days_raw}")
                recent_days = 365

        licensed_only = request.args.get('licensedOnly', 'false').lower() == 'true'
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'

        records = apply_disabled_filters(
            all_records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Last 365 Days"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('disabledDate', 'First Observed Disabled'),
            ('disabledDays', 'Days Since Observed Disabled'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'disabledDate' and value:
                    dt = parse_graph_datetime(value)
                    value = dt.date().isoformat() if dt else value
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-last-365-days-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting recently disabled report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/hired-this-year')
@require_auth
def get_recently_hired_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = load_recently_hired_data(report_cache, force_refresh=refresh)
        generated_at = None
        if os.path.exists(RECENTLY_HIRED_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(RECENTLY_HIRED_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at
        })
    except Exception as e:
        logger.error(f"Error loading recently hired report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/hired-this-year/export')
@require_auth
def export_recently_hired_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = load_recently_hired_data(report_cache, force_refresh=refresh)

        wb = Workbook()
        ws = wb.active
        ws.title = "Hired Last 365 Days"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('hireDate', 'Hire Date'),
            ('daysSinceHire', 'Days Since Hire'),
            ('managerName', 'Manager'),
            ('phone', 'Phone'),
            ('businessPhone', 'Business Phone'),
            ('location', 'Location')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'hireDate' and value:
                    value = format_hire_date(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"hired-last-365-days-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting recently hired report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


def _parse_bool_arg(value, default=True):
    if value is None:
        return default
    lowered = value.strip().lower()
    if lowered in {'true', '1', 'yes', 'on'}:
        return True
    if lowered in {'false', '0', 'no', 'off'}:
        return False
    return default


@app.route('/api/reports/last-logins')
@require_auth
def get_last_logins_report():
    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_never_signed_in = _parse_bool_arg(request.args.get('includeNeverSignedIn'), default=True)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=True)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=True)

        inactive_days_raw = request.args.get('inactiveDays')
        inactive_days = None
        if inactive_days_raw not in (None, '', 'null', 'None'):
            inactive_days = inactive_days_raw

        inactive_days_max_raw = request.args.get('inactiveDaysMax')
        inactive_days_max = None
        if inactive_days_max_raw not in (None, '', 'null', 'None'):
            inactive_days_max = inactive_days_max_raw

        records = load_last_login_data(report_cache, force_refresh=refresh)
        filtered_records = apply_last_login_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
            include_never_signed_in=include_never_signed_in,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            inactive_days=inactive_days,
            inactive_days_max=inactive_days_max
        )

        generated_at = None
        if os.path.exists(LAST_LOGIN_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(LAST_LOGIN_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': include_licensed and not include_unlicensed,
                'includeEnabled': include_enabled,
                'includeDisabled': include_disabled,
                'includeLicensed': include_licensed,
                'includeUnlicensed': include_unlicensed,
                'includeMembers': include_members,
                'includeGuests': include_guests,
                'includeNeverSignedIn': include_never_signed_in,
                'includeUserMailboxes': include_user_mailboxes,
                'includeSharedMailboxes': include_shared_mailboxes,
                'includeRoomEquipmentMailboxes': include_room_equipment_mailboxes,
                'inactiveDays': inactive_days,
                'inactiveDaysMax': inactive_days_max
            }
        })
    except Exception as error:
        logger.error(f"Error loading last sign-in report: {error}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/last-logins/export')
@require_auth
def export_last_logins_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_never_signed_in = _parse_bool_arg(request.args.get('includeNeverSignedIn'), default=True)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=True)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=True)
        inactive_days_raw = request.args.get('inactiveDays')
        inactive_days = None
        if inactive_days_raw not in (None, '', 'null', 'None'):
            inactive_days = inactive_days_raw

        inactive_days_max_raw = request.args.get('inactiveDaysMax')
        inactive_days_max = None
        if inactive_days_max_raw not in (None, '', 'null', 'None'):
            inactive_days_max = inactive_days_max_raw

        records = load_last_login_data(report_cache, force_refresh=refresh)
        filtered_records = apply_last_login_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
            include_never_signed_in=include_never_signed_in,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            inactive_days=inactive_days,
            inactive_days_max=inactive_days_max
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Users by Last Sign-In"

        headers = [
            ('name', 'Name'),
            ('title', 'Title'),
            ('department', 'Department'),
            ('email', 'Email'),
            ('accountEnabled', 'Account Enabled'),
            ('userType', 'User Type'),
            ('lastActivityDate', 'Most recent sign-in'),
            ('daysSinceLastActivity', 'Days since most recent sign-in'),
            ('lastInteractiveSignIn', 'Last interactive sign-in'),
            ('daysSinceInteractiveSignIn', 'Days since interactive sign-in'),
            ('lastNonInteractiveSignIn', 'Last non-interactive sign-in'),
            ('daysSinceNonInteractiveSignIn', 'Days since non-interactive sign-in'),
            ('neverSignedIn', 'Never signed in'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(filtered_records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'accountEnabled':
                    value = 'Yes' if record.get('accountEnabled', True) else 'No'
                elif key == 'userType':
                    user_type = (record.get('userType') or '').strip()
                    value = user_type.capitalize() if user_type else ''
                elif key == 'neverSignedIn':
                    value = 'Yes' if record.get('neverSignedIn') else 'No'
                elif key == 'licenseSkus' and isinstance(value, list):
                    value = ', '.join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 26

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"last-logins-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as error:
        logger.error(f"Error exporting last sign-in report: {error}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-licensed')
@require_auth
def get_disabled_licensed_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
        recent_days = base_filters.get('recentDays')
        filtered_records = apply_disabled_filters(
            all_records,
            licensed_only=True,
            recent_days=recent_days,
            include_guests=include_guests
        )
        generated_at = None
        if os.path.exists(DISABLED_LICENSE_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(DISABLED_LICENSE_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': True,
                'recentDays': recent_days,
                'includeGuests': include_guests
            }
        })
    except Exception as e:
        logger.error(f"Error loading disabled licensed report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-licensed/export')
@require_auth
def export_disabled_licensed_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )
        recent_days = base_filters.get('recentDays')
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
        records = apply_disabled_filters(
            all_records,
            licensed_only=True,
            recent_days=recent_days,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Licensed Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-licensed-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting disabled licensed report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/filtered-users')
@require_auth
def get_filtered_users_report():
    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=True)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=True)

        if 'licensedOnly' in request.args:
            legacy_licensed_only = _parse_bool_arg(request.args.get('licensedOnly'), default=True)
            if 'includeLicensed' not in request.args:
                include_licensed = True
            if 'includeUnlicensed' not in request.args:
                include_unlicensed = not legacy_licensed_only

        records = load_filtered_user_data(report_cache, force_refresh=refresh)
        filtered_records = apply_filtered_user_filters(
            records,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests
        )

        generated_at = None
        if os.path.exists(FILTERED_USERS_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(FILTERED_USERS_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'includeEnabled': include_enabled,
                'includeDisabled': include_disabled,
                'includeLicensed': include_licensed,
                'includeUnlicensed': include_unlicensed,
                'includeMembers': include_members,
                'includeGuests': include_guests,
                'includeUserMailboxes': include_user_mailboxes,
                'includeSharedMailboxes': include_shared_mailboxes,
                'includeRoomEquipmentMailboxes': include_room_equipment_mailboxes,
            }
        })
    except Exception as error:
        logger.error(f"Error loading filtered users report: {error}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/filtered-users/export')
@require_auth
def export_filtered_users_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_user_mailboxes = _parse_bool_arg(request.args.get('includeUserMailboxes'), default=True)
        include_shared_mailboxes = _parse_bool_arg(request.args.get('includeSharedMailboxes'), default=True)
        include_room_equipment_mailboxes = _parse_bool_arg(request.args.get('includeRoomEquipmentMailboxes'), default=True)

        if 'licensedOnly' in request.args:
            legacy_licensed_only = _parse_bool_arg(request.args.get('licensedOnly'), default=True)
            if 'includeLicensed' not in request.args:
                include_licensed = True
            if 'includeUnlicensed' not in request.args:
                include_unlicensed = not legacy_licensed_only

        records = load_filtered_user_data(report_cache, force_refresh=refresh)
        filtered_records = apply_filtered_user_filters(
            records,
            include_user_mailboxes=include_user_mailboxes,
            include_shared_mailboxes=include_shared_mailboxes,
            include_room_equipment_mailboxes=include_room_equipment_mailboxes,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('filterReasons', 'Filter Reasons'),
            ('accountEnabled', 'Account Enabled'),
            ('userType', 'User Type'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        reason_labels = {
            'filter_disabled': 'Hidden: disabled user',
            'filter_guest': 'Hidden: guest account',
            'filter_no_title': 'Hidden: missing title',
            'filter_ignored_title': 'Hidden: ignored title',
            'filter_ignored_department': 'Hidden: ignored department',
            'filter_ignored_employee': 'Hidden: ignored user'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(filtered_records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'filterReasons' and isinstance(value, list):
                    converted = [reason_labels.get(reason, reason) for reason in value]
                    value = ", ".join(converted)
                elif key == 'accountEnabled':
                    value = 'Yes' if record.get('accountEnabled', True) else 'No'
                elif key == 'userType':
                    user_type = (record.get('userType') or '').strip()
                    value = user_type.capitalize() if user_type else ''
                elif key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 26

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"filtered-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as error:
        logger.error(f"Error exporting filtered users report: {error}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/filtered-licensed')
@require_auth
def get_filtered_licensed_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = load_filtered_license_data(report_cache, force_refresh=refresh)
        generated_at = None
        if os.path.exists(FILTERED_LICENSE_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(FILTERED_LICENSE_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at
        })
    except Exception as e:
        logger.error(f"Error loading filtered licensed report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/filtered-licensed/export')
@require_auth
def export_filtered_licensed_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = load_filtered_license_data(report_cache, force_refresh=refresh)

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Licensed Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses'),
            ('filterReasons', 'Filter Reasons')
        ]

        reason_labels = {
            'filter_disabled': 'Hidden: disabled user',
            'filter_guest': 'Hidden: guest account',
            'filter_no_title': 'Hidden: missing title',
            'filter_ignored_title': 'Hidden: ignored title',
            'filter_ignored_department': 'Hidden: ignored department',
            'filter_ignored_employee': 'Hidden: ignored user'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'filterReasons' and isinstance(value, list):
                    converted = [reason_labels.get(reason, reason) for reason in value]
                    value = ", ".join(converted)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"filtered-licensed-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting filtered licensed report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500

@app.route('/api/auth-check')
def auth_check():
    """Simple endpoint to check if user is authenticated"""
    if session.get('authenticated'):
        return jsonify({'authenticated': True})
    else:
        return jsonify({'authenticated': False}), 401

@app.route('/api/search')
def search_employees():
    query = request.args.get('q', '').lower()
    
    if len(query) < 2:
        return jsonify([])
    
    try:
        if not os.path.exists(DATA_FILE):
            logger.warning(f"Data file {DATA_FILE} not found, attempting to fetch data")
            update_employee_data(source='search')
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
        else:
            logger.error("Could not create or find employee data file")
            return jsonify([])
        
        def flatten(node, results=None):
            if results is None:
                results = []
            if node and isinstance(node, dict):
                results.append(node)
                children = node.get('children', [])
                if children and isinstance(children, list):
                    for child in children:
                        flatten(child, results)
            return results
        
        all_employees = flatten(data)
        
        results = []
        for emp in all_employees:
            if emp and isinstance(emp, dict):
                name = emp.get('name') or ''
                title = emp.get('title') or ''
                department = emp.get('department') or ''
                
                name_match = query in name.lower()
                title_match = query in title.lower()
                dept_match = query in department.lower()
                
                if name_match or title_match or dept_match:
                    results.append(emp)
        
        return jsonify(results[:10])
    except FileNotFoundError as e:
        logger.error(f"File not found in search: {e}")
        return jsonify([])
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in search: {e}")
        return jsonify([])
    except AttributeError as e:
        logger.error(f"Attribute error in search (likely None value): {e}")
        logger.error(f"Query was: {query}")
        try:
            for emp in all_employees:
                if emp:
                    logger.debug(f"Employee data: name={emp.get('name')}, title={emp.get('title')}, dept={emp.get('department')}")
        except:
            pass
        return jsonify([])
    except Exception as e:
        logger.error(f"Error in search_employees: {e}")
        logger.error(f"Query was: {query}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/employee/<employee_id>')
def get_employee(employee_id):
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        def find_employee(node, target_id):
            if node.get('id') == target_id:
                return node
            for child in node.get('children', []):
                result = find_employee(child, target_id)
                if result:
                    return result
            return None
        
        employee = find_employee(data, employee_id)

        if employee:
            return jsonify(employee)
        return jsonify({'error': 'Employee not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update-now', methods=['POST'])
@require_auth
@limiter.limit(RATE_LIMIT_REFRESH)
def trigger_update():
    try:
        current_status = load_data_update_status()
        if current_status.get('state') == 'running':
            return jsonify({'error': 'Update already in progress'}), 409

        worker = threading.Thread(
            target=update_employee_data,
            kwargs={'source': 'manual'},
            daemon=True,
        )
        worker.start()
        logger.info(f"Manual update triggered by user: {session.get('username')}")
        return jsonify({'message': 'Update started'}), 200
    except Exception as e:
        logger.error(f"Error triggering update: {e}")
        mark_data_update_finished(success=False, error=str(e), source='manual')
        return jsonify({'error': 'Update failed'}), 500

@app.route('/search-test')
def search_test():
    return render_template_string(get_template('search_test.html'))

@app.route('/api/debug-search')
def debug_search():
    """Debug endpoint to check search functionality"""
    try:
        info = {
            'data_file_exists': os.path.exists(DATA_FILE),
            'data_file_path': os.path.abspath(DATA_FILE) if os.path.exists(DATA_FILE) else 'Not found',
            'data_file_size': os.path.getsize(DATA_FILE) if os.path.exists(DATA_FILE) else 0,
        }
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
                
                def count_employees(node):
                    count = 1
                    for child in node.get('children', []):
                        count += count_employees(child)
                    return count
                
                info['total_employees'] = count_employees(data) if data else 0
                info['root_employee'] = data.get('name', 'Unknown') if data else 'No data'
                info['has_children'] = bool(data.get('children')) if data else False
                
                def flatten(node, results=None):
                    if results is None:
                        results = []
                    if node and isinstance(node, dict):
                        results.append({
                            'id': node.get('id'),
                            'name': node.get('name'),
                            'title': node.get('title'),
                            'department': node.get('department')
                        })
                        children = node.get('children', [])
                        if children and isinstance(children, list):
                            for child in children:
                                flatten(child, results)
                    return results
                
                all_employees = flatten(data)
                info['sample_employees'] = all_employees[:5] if all_employees else []
                info['searchable_count'] = len(all_employees)
        else:
            info['error'] = 'Data file does not exist. Try triggering an update.'
            
        return jsonify(info)
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/force-update', methods=['POST'])
@require_auth
@limiter.limit(RATE_LIMIT_REFRESH)
def force_update():
    """Force an immediate update and wait for completion"""
    try:
        logger.info("Force update requested")
        update_employee_data(source='force-update')
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
                
            def count_employees(node):
                if not node:
                    return 0
                count = 1
                for child in node.get('children', []):
                    count += count_employees(child)
                return count
                
            total = count_employees(data)
            return jsonify({
                'success': True,
                'message': f'Data updated successfully. {total} employees in hierarchy.',
                'file_created': True
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Update completed but no data file created. Check Azure AD credentials.',
                'file_created': False
            })
    except Exception as e:
        logger.error(f"Force update error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=APP_PORT)